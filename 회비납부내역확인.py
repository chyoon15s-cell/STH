import streamlit as st
import pandas as pd
import os
import glob
from openpyxl import load_workbook

# 1. 안내 문구 설정
ERROR_MESSAGE = "오류입니다. 담당자에게 연락부탁드립니다. 070-4820-2709"
ELDERLY_NOTICE = "⚠️ 원로회원 변경 요청 문의필요 070-765-6503"

# 2. 페이지 설정
st.set_page_config(page_title="서울연극협회 회비 조회", layout="centered")

# 3. 데이터 로드 함수
@st.cache_data
def load_data_with_logic():
    try:
        excel_files = glob.glob("*.xlsx")
        if not excel_files: return None
        file_name = excel_files[0] 
        
        df = pd.read_excel(file_name, dtype=str)
        # 모든 열 이름의 줄바꿈과 공백을 제거하여 매칭 확률을 높입니다.
        df.columns = [str(c).replace('\n', '').strip() for c in df.columns]

        wb = load_workbook(file_name, data_only=True)
        ws = wb.active
        
        name_idx = -1
        elderly_col_idx = -1
        for i, cell in enumerate(ws[1]):
            header = str(cell.value).replace('\n', '').replace(' ', '').strip()
            if "성명" in header: name_idx = i + 1
            if "원로" in header: elderly_col_idx = i + 1

        elderly_target_rows = []
        for row in range(2, ws.max_row + 1):
            is_yellow = False
            if name_idx != -1:
                color = ws.cell(row=row, column=name_idx).fill.start_color.index
                if color in ["FFFF0000", "FFFFFF00", "FFFF00", "00FFFF00"]:
                    is_yellow = True
            
            has_elderly_text = False
            if elderly_col_idx != -1:
                val = str(ws.cell(row=row, column=elderly_col_idx).value).strip()
                if val and val not in ["None", "0", "nan"]:
                    has_elderly_text = True
            
            elderly_target_rows.append(is_yellow and has_elderly_text)
            
        df['is_elderly_target'] = elderly_target_rows[:len(df)]
        return df
    except:
        return None

df = load_data_with_logic()

# 4. 화면 구성
st.title("🎭 회비 납부 현황 조회")
st.write("성함과 생년월일 6자리를 입력해 주세요.")
st.markdown("---")

if df is None:
    st.error(ERROR_MESSAGE)
else:
    with st.form("search_form"):
        name_input = st.text_input("성함", placeholder="예: 홍길동")
        birth_input = st.text_input("생년월일 6자리", placeholder="예: 900101", max_chars=6)
        submit = st.form_submit_button("조회하기")

    if submit:
        if name_input and len(birth_input) == 6:
            try:
                # 데이터 검색
                match = df[
                    (df['성명'].str.replace(' ', '').str.strip() == name_input.replace(' ', '').strip()) & 
                    (df['생년월일'].str.contains(birth_input.strip()))
                ]
                
                if not match.empty:
                    res = match.iloc[0]
                    st.success(f"✅ {name_input} 회원님의 정보가 확인되었습니다.")
                    
                    if res['is_elderly_target'] == True:
                        st.warning(ELDERLY_NOTICE)

                    # 💥 수정된 미납 금액 로직: "2026년 기준 미납" 헤더를 직접 조준합니다.
                    fee_col = "2026년 기준 미납"
                    
                    if fee_col in df.columns:
                        raw_val = str(res[fee_col]).strip()
                        # 숫자만 남기고 정리
                        clean_val = raw_val.replace(',', '').replace('원', '').replace('.0', '')
                        
                        col1, col2 = st.columns(2)
                        
                        # 금액이 숫자로 있고 0보다 큰 경우 (미납)
                        if clean_val.isdigit() and int(clean_val) > 0:
                            with col1: st.metric("2026년 완납 여부", "🔴 미납")
                            with col2: st.metric("납부 예정 금액", f"{format(int(clean_val), ',')}원")
                        # 금액이 0이거나 데이터에 '완납'이라고 적힌 경우
                        elif clean_val == '0' or '완납' in raw_val:
                            with col1: st.metric("2026년 완납 여부", "🔵 완납")
                            with col2: st.metric("납부 예정 금액", "0원")
                        # 데이터가 nan이거나 비어있을 때
                        else:
                            with col1: st.metric("2026년 완납 여부", "🔴 미납")
                            with col2: st.metric("납부 예정 금액", "문의필요")
                    else:
                        st.warning(f"데이터에 '{fee_col}' 칸이 없습니다. 담당자에게 확인 부탁드립니다.")

                    # 소속 정보 (nan 제거)
                    def clean_info(val):
                        val = str(val).strip()
                        return "" if val.lower() in ['nan', 'none', ''] else val

                    branch = clean_info(res.get('소속지부', ''))
                    troupe = clean_info(res.get('소속극단', ''))
                    if branch or troupe:
                        st.info(f"**소속:** {branch} {'/' if branch and troupe else ''} {troupe}")
                    
                else:
                    st.warning("일치하는 정보가 없습니다. 입력 정보를 다시 확인해 주세요.")
            except:
                st.error(ERROR_MESSAGE)
        else:
            st.warning("성함과 생년월일 6자리를 올바르게 입력해 주세요.")
